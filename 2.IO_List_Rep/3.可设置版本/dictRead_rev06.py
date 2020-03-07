# -*- coding:utf-8 -*-
#!/usr/bin/python
#==========================================================
# Rev06
# 增加一个配置文件
# 报警弹窗
#==========================================================

import xlrd
import openpyxl
import os
import re
import sys
import logging
import openpyxl.styles as sty
import ConfigParser
import tkinter.messagebox 


#子程序=====================================================
def config():
    config_file_path = "config.ini"
    cf = ConfigParser.ConfigParser()
    cf.read(config_file_path)
    ExcleName = cf.get("baseconf", "ExcleName")
    newsheet = int(cf.get("baseconf", "newsheet"))
    oldsheet = int(cf.get("baseconf", "oldsheet"))
    config_list =[ExcleName,newsheet,oldsheet]
    return config_list

def rowList(sheet):
    rowList = []
    rowN=0
    rowsNum = sheet.nrows
    colNum = sheet.ncols
    while (rowN < rowsNum):
        row_Str = sheet.cell_value(rowN,2)
        rowList.append(row_Str)
        rowN=rowN+1
    ##print rowList
    return rowList

def colList(sheet):
    colList = []
    colN=0
    rowsNum = sheet.nrows
    colNum = sheet.ncols    
    while (colN < colNum):
        col_Str = sheet.cell_value(0,colN)
        colList.append(col_Str)
        colN=colN+1
    ##print rowList
    return colList
    
def dictGenerate(sheet):
    dict = {}
    rowsNum1 = sheet.nrows
    colNum1 = sheet.ncols
    for rowindex in range(rowsNum1):
        if rowindex != 0:
            for colindex in range(colNum1):
                if colindex > 2:
                    TAG_Str = sheet.cell_value(rowindex,2)
                    condition_Str = sheet.cell_value(0,colindex)
                    dict_Str = sheet.cell_value(rowindex,colindex)
                    dict[(TAG_Str,condition_Str)]=dict_Str                   
    return dict
    
def ExcelSave(filename,new,old):
    filePath = os.path.join(os.getcwd(), filename)
    excelFile = xlrd.open_workbook(filePath)
    #print excelFile.sheet_names() 
    sheet1 = excelFile.sheet_by_index(new) 
    sheet2 = excelFile.sheet_by_index(old) 

    #==========================================================
    rowsNum = sheet1.nrows
    colNum = sheet1.ncols
    #print "row num:", sheet1.nrows #行
    #print "col num:", sheet1.ncols #列
    rowlist_list=rowList(sheet1)
    collist_list=colList(sheet1)   
    dictTAG = dictGenerate(sheet1)
   
    #读取更换内容==============================================
    newTAG = dictGenerate(sheet2)
    list_TAG = list(newTAG)
    #print list_TAG # [(u'BC', u'AB')]
    TAG_index=0
    #openpyxl 打开Yget 清单=============================================
    data = openpyxl.load_workbook(filename)
    table = data.worksheets[new]
    
    while(TAG_index < len(list_TAG)):
        TAG_XY = list_TAG[TAG_index]
        #print TAG_XY
        TAG_Z = newTAG[TAG_XY]
        #print TAG_Z # 7.0
        Nokey = dictTAG.get(TAG_XY,0)  #处理Yget 不存在的key
        
        if (TAG_Z !="" and Nokey !=0):
            print TAG_Z
            #Excel坐标============================================
            NewTagRow_Str = list_TAG[TAG_index][0]
            NewTagcol_Str = list_TAG[TAG_index][1]
            NewTagRow_index = rowlist_list.index(NewTagRow_Str)+1
            NewTagCel_index = collist_list.index(NewTagcol_Str)+1
            #替换Excel数据===============================================
            table.cell(NewTagRow_index,NewTagCel_index).value = TAG_Z 
            table.cell(NewTagRow_index,NewTagCel_index).fill=sty.PatternFill(fill_type='solid',fgColor="00FFFF") #对更新数据进行标注颜色
        TAG_index= TAG_index+1 
    #Excel写入===========================================================    
    data.save(filename)



#主程序====================================================== 
if __name__ == "__main__":  
    try:
        config = config()
        filename = config[0]
        new_index = config[1]-1
        old_index = config[2]-1
        ExcelSave(filename,new_index,old_index)
    except IOError,e:
        tkinter.messagebox.showinfo('提示','请勿打开Excle文档')
    except ConfigParser.NoSectionError,e:
        tkinter.messagebox.showinfo('提示','配置文件丢失' )
    except IndexError,e:
        tkinter.messagebox.showinfo('提示','配置文件设置错误' )
        
