# -*- coding:utf-8 -*-
#!/usr/bin/python
# python 3.7
#==========================================================

# Rev01,Rev02
# 可以读取Yget 清单，并使用第四个表 的内容替换 原来清单数据 
# 只能更换一个数据

# Rev03
# 替换 多个数据

# Rev05
# 第3张表数据 更新第2张表内容
# 配合表格 根据Yget清单跑机柜接线表

# Rev06
# 增加一个配置文件
# 报警弹窗

# Rev07
# 增加 校队功能

# Rev08
# 增加更多参数
# 增加文档信息确认的弹窗
#==========================================================

import xlrd
import openpyxl
import os
import re
import sys
import logging
import openpyxl.styles as sty
import ConfigParser
import codecs
import tkinter.messagebox 
#from tkinter import *


#子程序=====================================================
def config():
    config_file_path = "config.ini"
    cf = ConfigParser.ConfigParser()
    cf.read(config_file_path, encoding ='UTF-8-SIG')
    #encoding='UTF-8'



    mode = int(cf.get("baseconf", "mode"))
    
    TExcleName = cf.get("TargetExcel", "ExcleName")
    TSheetNum = int(cf.get("TargetExcel", "SheetNum"))
    TtagColNum = int(cf.get("TargetExcel", "TagColNum"))
    
    RExcleName = cf.get("ReferExcel", "ExcleName")
    RSheetNum = int(cf.get("ReferExcel", "SheetNum"))
    RtagColNum = int(cf.get("ReferExcel", "TagColNum"))

    config_list =[mode,TExcleName,TSheetNum,TtagColNum,RExcleName,RSheetNum,RtagColNum]
    return config_list

def rowList(sheet,Cindex):
    rowList = []
    rowN=0
    rowsNum = sheet.nrows
    colNum = sheet.ncols
    while (rowN < rowsNum):
        row_Str = sheet.cell_value(rowN,Cindex)
        rowList.append(row_Str)
        rowN=rowN+1
    ##print rowList
    return rowList

def colList(sheet):
    colList = []
    colN=0
    rowsNum = sheet.nrows
    colNum = sheet.ncols    
    while (colN < colNum):
        col_Str = sheet.cell_value(0,colN)
        colList.append(col_Str)
        colN=colN+1
    ##print rowList
    return colList
    
def dictGenerate(sheet,Cindex):
    dict = {}
    rowsNum1 = sheet.nrows
    colNum1 = sheet.ncols
    for rowindex in range(rowsNum1):
        if rowindex != 0:
            for colindex in range(colNum1):
                if colindex > Cindex:
                    TAG_Str = sheet.cell_value(rowindex,Cindex)
                    condition_Str = sheet.cell_value(0,colindex)
                    dict_Str = sheet.cell_value(rowindex,colindex)
                    dict[(TAG_Str,condition_Str)]=dict_Str                   
    return dict
    
def ExcelSave(*config):    #替换功能

    TExcleName = config[1]
    TSheetNum = config[2]-1
    TtagColNum = config[3]-1
    RExcleName = config[4]
    RSheetNum = config[5]-1
    RtagColNum = config[6]-1 
    
    
    TfilePath = os.path.join(os.getcwd(), TExcleName)
    TexcelFile = xlrd.open_workbook(TfilePath)
    RfilePath = os.path.join(os.getcwd(), RExcleName)
    RexcelFile = xlrd.open_workbook(RfilePath)
    #print excelFile.sheet_names() 
    
    sheet1 = TexcelFile.sheet_by_index(TSheetNum) 
    sheet2 = RexcelFile.sheet_by_index(RSheetNum) 

    #==========================================================
    rowsNum = sheet1.nrows
    colNum = sheet1.ncols
    #print "row num:", sheet1.nrows #行
    #print "col num:", sheet1.ncols #列
    rowlist_list=rowList(sheet1,TtagColNum)
    collist_list=colList(sheet1)   
    dictTAG = dictGenerate(sheet1,TtagColNum)
   
    #读取更换内容==============================================
    newTAG = dictGenerate(sheet2,RtagColNum)
    list_TAG = list(newTAG)
    #print list_TAG # [(u'BC', u'AB')]
    TAG_index=0
    #openpyxl 打开Yget 清单=============================================
    data = openpyxl.load_workbook(TExcleName)
    table = data.worksheets[TSheetNum]
    
    while(TAG_index < len(list_TAG)):
        TAG_XY = list_TAG[TAG_index]
        #print TAG_XY
        TAG_Z = newTAG[TAG_XY]
        #print TAG_Z # 7.0
        Nokey = dictTAG.get(TAG_XY,0)  #处理Yget 不存在的key
        
        if (TAG_Z !="" and Nokey !=0):
            print(TAG_Z)
            #Excel坐标============================================
            NewTagRow_Str = list_TAG[TAG_index][0]
            NewTagcol_Str = list_TAG[TAG_index][1]
            NewTagRow_index = rowlist_list.index(NewTagRow_Str)+1
            NewTagCel_index = collist_list.index(NewTagcol_Str)+1
            #替换Excel数据===============================================
            table.cell(NewTagRow_index,NewTagCel_index).value = TAG_Z 
            table.cell(NewTagRow_index,NewTagCel_index).fill=sty.PatternFill(fill_type='solid',fgColor="00FFFF") #对更新数据进行标注颜色  
        TAG_index= TAG_index+1 
    #Excel写入===========================================================    
    data.save(TExcleName)

def ExcelComp(*config):    #比较功能
    TExcleName = config[1]
    TSheetNum = config[2]-1
    TtagColNum = config[3]-1
    RExcleName = config[4]
    RSheetNum = config[5]-1
    RtagColNum = config[6]-1 

    TfilePath = os.path.join(os.getcwd(), TExcleName)
    TexcelFile = xlrd.open_workbook(TfilePath)
    RfilePath = os.path.join(os.getcwd(), RExcleName)
    RexcelFile = xlrd.open_workbook(RfilePath)
    #print excelFile.sheet_names() 
    
    sheet1 = TexcelFile.sheet_by_index(TSheetNum) 
    sheet2 = RexcelFile.sheet_by_index(RSheetNum) 

    #==========================================================
    rowsNum = sheet1.nrows
    colNum = sheet1.ncols
    #print "row num:", sheet1.nrows #行
    #print "col num:", sheet1.ncols #列
    rowlist_list=rowList(sheet1,TtagColNum)
    collist_list=colList(sheet1)   
    dictTAG = dictGenerate(sheet1,TtagColNum)
   
    #读取更换内容==============================================
    newTAG = dictGenerate(sheet2,RtagColNum)    
    list_TAG = list(newTAG)
    #print list_TAG # [(u'BC', u'AB')]
    TAG_index=0
    #openpyxl 打开Yget 清单=============================================
    data = openpyxl.load_workbook(TExcleName)
    table = data.worksheets[TSheetNum]
    
    while(TAG_index < len(list_TAG)):
        TAG_XY = list_TAG[TAG_index]
        #print TAG_XY
        TAG_Z = newTAG[TAG_XY]
        #print TAG_Z # 7.0
        Nokey = dictTAG.get(TAG_XY,0)  #处理Yget 不存在的key 
        if (TAG_Z !="" and Nokey !=0):
            
            #Excel坐标============================================
            NewTagRow_Str = list_TAG[TAG_index][0]
            NewTagcol_Str = list_TAG[TAG_index][1]
            NewTagRow_index = rowlist_list.index(NewTagRow_Str)+1
            NewTagCel_index = collist_list.index(NewTagcol_Str)+1
            #替换Excel数据===============================================
            TAG_ZZ = table.cell(NewTagRow_index,NewTagCel_index).value
            #table.cell(NewTagRow_index,NewTagCel_index).value = TAG_Z 
            if (TAG_Z != TAG_ZZ):
                print (TAG_Z)
                table.cell(NewTagRow_index,NewTagCel_index).fill=sty.PatternFill(fill_type='solid',fgColor="FF00FF") #对更新数据进行标注颜色  
        TAG_index= TAG_index+1 
    #Excel写入===========================================================    
    data.save(TExcleName)

def reading(*config):
    mode = config[0]
    TExcleName = config[1]
    TSheetNum = config[2]-1
    TtagColNum = config[3]-1
    RExcleName = config[4]
    RSheetNum = config[5]-1
    RtagColNum = config[6]-1 
    TfilePath = os.path.join(os.getcwd(), TExcleName)
    TexcelFile = xlrd.open_workbook(TfilePath)
    RfilePath = os.path.join(os.getcwd(), RExcleName)
    RexcelFile = xlrd.open_workbook(RfilePath)
    if mode == 1:
        modeStr ="替换模式"
    elif mode == 2:
        modeStr ="校队模式"
    else:
        modeStr ="配置文件设置错误"

    #=========================
    label1 = '请确认以下信息：'
    label2 = '目标文档：'+ TfilePath
    label3 = '目标Sheet：'+ str(TexcelFile.sheet_names()[TSheetNum])
    label4 = '参考文档：'+ RfilePath  
    label5 = '参考Sheet：'+ str(RexcelFile.sheet_names()[RSheetNum])
    label6 = '模式：'+ modeStr
    label = label1+'\n'+label2+'\n'+label3+'\n'+label4+'\n'+label5+'\n'+label6
    #label = label.encode('gb2312')
    comfirm = tkinter.messagebox.askquestion(title='文档信息',message=label)  
    return comfirm
 
#主程序====================================================== 
if __name__ == "__main__":     
    try:
        #配置文件读取======================================================
        config = config()
        mode = config[0]
        comfirm = reading(*config)
        if comfirm == "yes":
            #模式选择======================================================
            if mode == 1:
                ExcelSave(*config)
            elif mode == 2:
                ExcelComp(*config)
            else:
                tkinter.messagebox.showinfo('提示','配置文件设置错误' )
                #tkinter.messagebox.askquestion(title='hello',message='hello word')   

    #错误输出================================================
    except IOError as e:
        tkinter.messagebox.showinfo('提示','请勿打开Excle文档')
    except ConfigParser.NoSectionError as e:
        tkinter.messagebox.showinfo('提示','配置文件丢失' )
    except IndexError as e:
        tkinter.messagebox.showinfo('提示','配置文件设置错误' )
        
