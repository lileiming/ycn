# -*- coding:utf-8 -*-
#!/usr/bin/python
#==========================================================
# Rev03
# 替换 多个数据
#==========================================================

import xlrd
import openpyxl
import os
import re
import sys
import logging
import openpyxl.styles as sty

#变量======================================================
#rowList = []
#colList = []


#Excel读取==================================================
filename = "Yget.xlsx"
filePath = os.path.join(os.getcwd(), filename)
excelFile = xlrd.open_workbook(filePath)
#print excelFile.sheet_names() 

sheet1 = excelFile.sheet_by_index(2) 
sheet2 = excelFile.sheet_by_index(3) 

#==========================================================
rowsNum = sheet1.nrows
colNum = sheet1.ncols
#print "row num:", sheet1.nrows #行
#print "col num:", sheet1.ncols #列
    
#子程序=====================================================
def rowList(sheet1):
    rowList = []
    rowN=0
    while (rowN < rowsNum):
        row_Str = sheet1.cell_value(rowN,2)
        rowList.append(row_Str)
        rowN=rowN+1
    ##print rowList
    return rowList

def colList(sheet1):
    colList = []
    colN=0
    while (colN < colNum):
        col_Str = sheet1.cell_value(0,colN)
        colList.append(col_Str)
        colN=colN+1
    ##print rowList
    return colList
    
def dictGenerate(sheet):
    dict = {}
    rowsNum1 = sheet.nrows
    colNum1 = sheet.ncols
    for rowindex in range(rowsNum1):
        if rowindex != 0:
            for colindex in range(colNum1):
                if colindex > 2:
                    TAG_Str = sheet.cell_value(rowindex,2)
                    condition_Str = sheet.cell_value(0,colindex)
                    dict_Str = sheet.cell_value(rowindex,colindex)
                    dict[(TAG_Str,condition_Str)]=dict_Str                   
    return dict

#主程序====================================================== 
if __name__ == "__main__":  
    rowlist_list=rowList(sheet1)
    collist_list=colList(sheet1)   
    dictTAG = dictGenerate(sheet1)
   
    #读取更换内容==============================================
    newTAG = dictGenerate(sheet2)
    list_TAG = list(newTAG)
    #print list_TAG # [(u'BC', u'AB')]
    TAG_index=0
    #openpyxl 打开Yget 清单=============================================
    data = openpyxl.load_workbook(filename)
    table = data.worksheets[2]
    
    while(TAG_index < len(list_TAG)):
        TAG_XY = list_TAG[TAG_index]
        #print TAG_XY
        TAG_Z = newTAG[TAG_XY]
        print TAG_Z # 7.0
        Nokey = dictTAG.get(TAG_XY,0)  #处理Yget 不存在的key
        
        if (TAG_Z !="" and Nokey !=0):
            #Excel坐标============================================
            NewTagRow_Str = list_TAG[TAG_index][0]
            NewTagcol_Str = list_TAG[TAG_index][1]
            NewTagRow_index = rowlist_list.index(NewTagRow_Str)+1
            NewTagCel_index = collist_list.index(NewTagcol_Str)+1
            #替换Excel数据===============================================
            table.cell(NewTagRow_index,NewTagCel_index).value = TAG_Z 
            table.cell(NewTagRow_index,NewTagCel_index).fill=sty.PatternFill(fill_type='solid',fgColor="00FFFF") #对更新数据进行标注颜色
        TAG_index= TAG_index+1 
    #Excel写入===========================================================    
    data.save(filename)

