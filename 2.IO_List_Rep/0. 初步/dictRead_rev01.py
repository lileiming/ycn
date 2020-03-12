# -*- coding:utf-8 -*-
#!/usr/bin/python

import xlrd
import openpyxl
import os
import re
import sys
import logging

#变量======================================================
#rowList = []
#colList = []


#Excel读取==================================================
filename = "1A.xlsx"
filePath = os.path.join(os.getcwd(), filename)
excelFile = xlrd.open_workbook(filePath)
print excelFile.sheet_names() 

#sheet1 = excelFile.sheet_by_index(3) 
sheet1 = excelFile.sheet_by_name(u'Sheet1')
sheet2 = excelFile.sheet_by_name(u'Sheet2')

#==========================================================
rowsNum = sheet1.nrows
colNum = sheet1.ncols
#print "row num:", sheet1.nrows #行
#print "col num:", sheet1.ncols #列
    
#子程序=====================================================
def rowList(sheet1):
    rowList = []
    rowN=0
    while (rowN < rowsNum):
        row_Str = sheet1.cell_value(rowN,0)
        rowList.append(row_Str)
        rowN=rowN+1
    ##print rowList
    return rowList

def colList(sheet1):
    colList = []
    colN=0
    while (colN < colNum):
        col_Str = sheet1.cell_value(0,colN)
        colList.append(col_Str)
        colN=colN+1
    ##print rowList
    return colList


def dictGenerate(sheet):
    dict = {}
    rowsNum1 = sheet.nrows
    colNum1 = sheet.ncols
    for rowindex in range(rowsNum1):
        if rowindex != 0:
            for colindex in range(colNum1):
                if colindex != 0:
                    TAG_Str = sheet.cell_value(rowindex,0)
                    condition_Str = sheet.cell_value(0,colindex)
                    dict_Str = sheet.cell_value(rowindex,colindex)
                    dict[(TAG_Str,condition_Str)]=dict_Str                   
    return dict

#主程序====================================================== 
if __name__ == "__main__":  
    rowlist_list=rowList(sheet1)
    collist_list=colList(sheet1)    
    dictTAG = dictGenerate(sheet1)
    #print(u"最终获取到的数据是：{0}".format(dictTAG))   
    #print dictTAG.get((u'BC', u'AC'),u"没有这个Key！")
    
    #替换字典内容==============================================
    newTAG = {(u'BC', u'AB'): 7.0}
    #newTAG = dictGenerate(sheet2)
    #print newTAG
    list_TAG = list(newTAG)
    #print list_TAG # [(u'BC', u'AB')]
    TAG_XY = list_TAG[0]
    #print TAG_XY
    TAG_Z = newTAG[TAG_XY]
    #print TAG_Z # 7.0
    dictTAG[TAG_XY] = TAG_Z
    #print(u"最终获取到的数据是：{0}".format(dictTAG)) 

    #Excel坐标============================================
    NewTagRow_Str = list_TAG[0][0]
    NewTagcol_Str = list_TAG[0][1]
    NewTagRow_index = rowlist_list.index(NewTagRow_Str)+1
    NewTagCel_index = collist_list.index(NewTagcol_Str)+1
    #print (NewTagRow_index,NewTagCel_index)
    
    #Excel写入===========================================================
    
    data = openpyxl.load_workbook('1A.xlsx')
        #print(data.get_named_ranges()) # 输出工作页索引范围
        #print(data.get_sheet_names()) # 输出所有工作页的名称
        #sheetnames = data.get_sheet_names()
        #table = data.get_sheet_by_name(sheetnames[1])
    table = data["Sheet1"]
        #table = data.active #当前表
        #print(table.title) # 输出表名
        #nrows = table.max_row # 获得行数
        #ncolumns = table.max_column # 获得行数
    table.cell(NewTagRow_index,NewTagCel_index).value = TAG_Z
        
    data.save('1A.xlsx')
    
    
    
    
    
    
    