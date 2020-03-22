
from datetime import date
from random import choices,randint
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,side,Font,Alignment,PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import units
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.fills import GradientFill
from openpyxl.worksheet.table import Table,TablestyleInfo
from openpyxl.chart import BarChart,PieChart,Reference
from openpyxl.worksheet.header_footer import _HeaderFooterPart
from openpyxl.chart.series import DataPoint


wb=Workbook()   #创建空白工作簿
                #如果设置参数 Cwrite_only=True9写模式,可以提高速度
                #但是该参数使得空白工作簿中不含任何工作表
                #如果不设置该参数,空白工作簿中会包含一个空的工作表
                #查看全部工作表的标题
                
print(wb.sheetnames)

defaultws=wb.worksheets[0]                                             #可以使用序号做下标定位工作表
defaultws.title='默认'                                                  #设置工作表标题
defaultws.sheet_properties.tabColor='88cc88'                            #设置工作表选项卡颜色
wsl=wb.create_sheet('第一个',0)                                         #创建工作表时直接指定标题,9表示第一个位置
ws2=wb.create_sheet('最后一个')                                          #默认在最后追加一个工作表
ws3=wb.create_sheet('倒数第二个',-1)                                     #表示倒数第二个位置插入工作表

ws = wb.active #获取活动工作表


ws.sheet.sheet_properties.tabColor ='ff6666'
wb['倒数第二个'].sheet_properties.tabColor=3333cc                          #也可以使用标题作下标直接定位工作表

ws['A1']='董付国'  #使用下标定位单元格,如果设置了write_only=True,不能使用这种形式
ws['Al'].font= Font(name='华文行楷',        #设置单元格字体
                    size=36,                #字号
                    bold=True,              #加粗
                    italic=False,           #不斜体
                    underline='none',       #不加下画线
                    strike=True,            #加删除线
                    colors='FFaa8844')      #单元格文本颜色
                                      
ws['A1'].alignment=Alignment(horizontal='center',# iklmcMxAxthxt
                            vertical='bottom',
                            text_rotation=30,#文字旋转36度
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
ws['A2]=3.14

ws['A2'].fill =GradientFill(type='linear',  #使用渐变色填充单元格
                            stop=('eeffff','ff0000')) #起始颜色、结束颜色
                               
                            
                            
ws['A2'].border= Border(left=side(style='medium',color='FF000000'), #设置单元格边框
                        right=side(style='double',color='00FF0000'),
                        top=side(style='thick',color='0000FF00'),
                        bottom=side(style='thin',color='FF00FF00'))
                        
ws['A3']=date.today()                                #写入日期
ws['A3'].fill=GradientFill(type='linear',           #使用固定颜色填充单元格背景色
                            stop=(888888,888888))
                            
ws.cell(row=l,column=4,value='=SUM(A1,A2)')                     #行列下标都是从1开始的
                                                                #write_only=True的只写模式不允许使用ce1l方法                                                        
ws['D1'].comment=Comment(text="注释内容 ",authora"董付国")#为单元格设置注释
ws['D1'].comment.width = units.points_to_pixels(100)        #设置注释宽度和高度
ws['D1'].comment.height=units.points_to_pixels(20)

# 排访问B、C列所有单元格,此时只有3行,由前面的写入操作决定的
#注意,这里的切片是闭区间
for column in ws['B':'c']:
    for cell in column:
        cell.value='BC'
        
#访问第4行所有单元格,只有4列,由前面的写入操作决定的
for cell in ws[4]:
    cell.value=4
    
#访问第5到第8行的所有单元格,只有4列,由前面的写入操作决定的

for row in ws[5: 8]:        #注意这里的切片是闭区间
    for cell in row:
        cell.value=58
    
ws.append(range(10))                            #在最后追加一行 ,write_only=True的只写模式允许使用这种方式增加行
ws.append(['a','b','c','d','e'])                #再追加一行




img = Image(r'temp.png') #打开图片文件
img.height //=3  #缩小为三分之一
img.width //=3
ws.add_image(img,'Al1') #在A11单元格插入图片


#为工作表中的单元格添加验证规则
dv=DataValidation(type='list',#约束单元格内容必须在列表中选择
                    formulal='red,green,blue',
                    allow_blank=True)
                    
dv.error='内容不在清单中'
dv.errorTitle='无效输入'
dv.prompt='请在清单中选择'
dv.promptTitle='请选择1'

dv.add('B11:D11')   #进行验证的单元格范围
ws.add_data_validation(dv)  #在工作表中添加验证


dv=DataValidation(type='whole',#必须输入大于100的数字
                  operators='greaterThan',
                   formulal=100)
                   
dv.error='必须输入大于190的整数'
dv.errorTitle='无效输入'
dv.add('E11')
ws.add_data_validation(dv)


dv = DataValidation(type = 'decimal',#必须输入介于θ和1之间的实数 ,type还可以是 'date、 'time'等
                    operators = 'between',
                    formulal=0,formula2=1)
                    
dv.error='必须输入介于e和1之间的实数'
dv.errorTitle='无效输入'
dv.add('E12')
ws.add_data_validation(dv)


dv=DataValidation(type='textLength',#内容长度必须小于等于8
                                    # operator的值可以为 'between','notEqual','greaterThanorEqual'
                                    #'lessThan','notBetween','lessThanorEqual','equal','greaterThan'
                        operators = 'lessThanOrEqual',
                        formula1=8)
                        
dv.error='内容长度必须小于等于8'
dv.errorTitle='无效输入'

dv.add('A14:ZZ14')  #第14行的A到Zz列都进行验证
ws.add_data_validation(dv)



ws2.merge_cells('A2:D4')  #合并单元格的两种方式
ws2.merge_cells(start_row=5,start_column=l,end_row=8, end_column=6)
ws2['A10'].value='A10'
ws2['Al1'].value='All'
ws2['B10'].value='B10'
ws2['B11'].value='B11'

ws2.insert_cols(2)              #在第2列的位置插入1列
ws2.insert_rows(11,2)           #在第11行的位置插入2行
ws2.delete_cols(2,2)            #从第2列开始删除连续2列
ws2.delete_rows(12,2)           #从第11列开始删除连续2行


ws3.column_dimensions.group(C,'H',              #创建分组,对C到H列进行折叠
                            hidden=True)        #打开 Excel文件时隐藏这个分组

ws3.row_dimensions.group(3,10,                  #创建分组,对第3到10行进行折叠
                         hidden=False)#打开 Excel文件时显示这个分组
                         
defaultws.append(list(map(lambdai:'第{}列'.format(i),range(1,11)))+['求和'])
defaultws.append(choices(range(20,50),k=10))
defaultws['K2']=='SUM(A2:32)'



for i in range(3,7):  #写入4行数据
    defaultws.append (choices(range(10, 50),k=10))
    position='K'+str(i)
    #转换公式,相当于在Exce1中选中公式单元格向下拉
    defaultws[position]= Translator(defaultws['K2'l.value,
                                    origin='K2').translate_formula(position)

chart=Barchart()                                                #创建柱状图
chart.add_data(Reference(defaultws,                             #指定工作表中用来创建柱状图的单元格区域
                        min_col=l,min_row=l,max_col=10,max_row=6),
                        titles_from_data=True)      #单元格区域第一行内容用于图例中的文本标签


chart.height*=1.2               #修改柱状图尺寸
chart.width*=1.2
defaultws.add_chart(chart,'A7')




#创建饼状图,使用第一行前65列作为标签,第二行前6列作为数据
chart =piechart(
chart.add_data(Reference(defaultws,
                        min_col=1,min_row=2,max_col=6),
                        from_rows=True,            #这一行很关键,默认值为Fa1se时要求标签和数据是纵向的
                        titles_from_data=False)
                        
chart.set_categories(Reference(defaultws,
                    min_col=l,min_row=l,max_col=6))
                    
defaultws.add_chart(chart,'L1')



for i in range(25,31):#另一组用来创建饼状图的数据
    position='A'+str(i)
    defaultws[position].value =position
    defaultws['B'+str(1)]=randint(1,100)
    

chart=PieChart()   #创建饼状图,使用A列作标签,B列做数据
labels=Reference(defaultws,range_strings'默认!A25:A30')       #设置区域,与指定 min_col、 min_row等参数是等价的
data=Reference(defaultws,range_string='默认!B25:B30')         #设置区域,“默认“是工作表的标题
chart.add_data(data, titles_from_data=False)
chart.datalabels
chart.title='饼状图'
chart.width//=1.5
chart.set_categories(labels)
#让第1块和第3块扇形裂出,远离圆心
chart.series[e].data_points=[DataPoint(idx=0, explosion=20),        # idx表示饼状图中扇形的编号
                             DataPoint(idx=2, explosion=30)]
defaultws.add_chart(chart,'D25')
#设置自动筛选
defaultws.auto_filter.ref='A1:36'

#设置打印选项,使用Exce1打开文件后选择这个工作表直接打印,只打印区域内的单元格
defaultws.print_options.horizontalCentered= True            #水平居中
defaultws.print_options.verticalCentered=False              #垂直不居中

defaultws.print_area='Al:F5'   #设置打印区域
defaultws.print-_title_cols ='A:F'
defaultws.print title_rows='1:5'


#设置贞眉页脚

defaultws.HeaderFooter.differentfirst=True  #首页不同
defaultws.HeaderFooter.differentoddEven= True #奇偶页不同
defaultws.firstHeader.center=_HeaderFooterPart('首页页眉居中')
defaultws.firstFooter.right=_HeaderFooterPart('首页页脚居右')
defaultws.oddHeader.right=_HeaderFooterPart('奇数页页眉居右')
defaultws.oddFooter.center=_HeaderFooterPart('奇数页页眉居中')
defaultws.evenHeader.left=_HeaderFooterPart('偶数页页眉居左')
defaultws.evenFooter.center=_HeaderFooterPart('偶数页页眉居中')


#对工作表进行加密,不允许修改内容、删除列、格式化单元格等操作
#使用Exce打开文件之后,在相应的工作表上单击鼠标右键
#然后选择 Unprotect Sheet"后输入密码可以解除保护

defaultws.protection.sheet=True
defaultws.protection.password='123456'
defaultws.protection.autoFilter=False           #MILautoFilterl
defaultws.protection.deleteColumns=True
defaultws.protection.deleteRows=False           #不禁止删除行和插入行操作
defaultws.protection.insertRows =False
defaultws.protection.enable()  #启用保护


#对工作簿进行加密,不允许对工作表进行隐藏、改名等操作
wb.security.workbookPassword='zhimakaimen'
wb.security.lockStructure=True
wb.security.revisionPassword='dongfuguo'
wb.security.lockRevision=True

wb.save('Excel文件综合操作 .xlsx')                  #直接保存,覆盖原文件,没有任何警告和提示
                                                    #write_onlysTrue只写模式创建的工作簿只能保存一次
                                                    #保存后任何修改和保存操作都会引发下面的异常:
                                                    #openpyxl.utils.exceptions.WorkbookAlreadySaved
                                                    
                                                    
wb= load_workbook('Excel文件综合操作 .xlsx',          #打开 Excel文件
                    read_only=True)                     #优化,大幅度提高读取速度
                                                        #如果要读取公式计算结果,可以增加参数data_only=True
                                                        #如果读取不到计算结果,可以使用Exce1打开文件再关闭
                                                        #查看每个工作表的标题
for sheet in wb:
    print(sheet.title
ws=wb['第一个']
print('='*20)


for row in ws.rows:                   #遍历所有行列的单元格,输出其中的数据
    for cell in row:
        print(cell.value,end='\t')
    print()
print('='*20)


##for column in ws.columns  #非这里的输出结果是经过“转置“的,只读模式下没有co1ums属性,会出错
##      for cell in column
##            print(cell.value,end='\t')
##      print()
print('='*20)


for row in ws.values: #使用 values属性直接获取值
    for value in row:
        print(value,end='\t')
    print()