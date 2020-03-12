# -*- coding:utf-8 -*-
#!/usr/bin/python
# python 3.7
#==========================================================

# Rev01
# 读取模板

#==========================================================

import openpyxl
from openpyxl  import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill


#子程序=====================================================

class App:
    def __init__(self):
        self.replace()

    def sampleAddr(self):   #Sample address
        #self.sampleNo =3
        #self.Redu = 0
        self.MinCol = 3
        self.MaxCol = 3 + self.Redu
        self.MinRow = 2 + (self.sampleNo-1)*37
        self.MaxRow = 37 + (self.sampleNo-1)*37
        
    def targetaddr(self):#target address
        #self.NodeNo = 2
        #self.SlotNo = 4
        if(self.NodeNo<6):
            self.MbMinCol = self.SlotNo+2
            self.MbMaxCol = self.SlotNo+2+self.Redu
            self.MbMinRow = 3+(self.NodeNo-1)*40
            self.MbMaxRow = 38+(self.NodeNo-1)*40
        elif(self.NodeNo>=6):
            self.MbMinCol = self.SlotNo+14
            self.MbMaxCol = self.SlotNo+14+self.Redu
            self.MbMinRow = 3+(self.NodeNo-6)*40
            self.MbMaxRow = 38+(self.NodeNo-6)*40
        
    def readExcel(self):
    
        self.wb = load_workbook('NodePlan.xlsx')
        self.ws = self.wb["Sample"]
        self.ws_charge = self.wb["卡件IO布置"]
        self.wskj = self.wb["卡件布置图"]
        
    def carddict(self):
        cdict = {'AAI143/R': 1,
                'AAI543/R': 2,
                'AAI143': 3,
                'ADV151/R':4,
                'ADV151':5,
                'ADV551/R':6,
                'ADV551':7,
                'ADCV01':8,
                'EC401':0,
                'R':0}
                
        redudict ={'AAI143/R': 1,
                'AAI543/R': 1,
                'AAI143': 0,
                'ADV151/R':1,
                'ADV151':0,
                'ADV551/R':1,
                'ADV551':0,
                'ADCV01':0,
                'EC401':0,
                'R':0}
        try:
            self.sampleNo = cdict[str(self.Cardtype)]
            self.Redu = redudict[str(self.Cardtype)]
        except KeyError as e:
            self.sampleNo = 9
            self.Redu = 0

    def replace(self):
        self.readExcel() 
        for kjcol in  self.wskj.iter_cols(max_col=62,max_row=95,min_col=62,min_row=16):
            for kjcell in  kjcol:
                self.NodeNo = kjcell.value
                #print(self.NodeNo)
                self.SlotNo = self.wskj.cell(kjcell.row,kjcell.column+1).value
                self.Cardtype = self.wskj.cell(kjcell.row,kjcell.column+2).value
                #print(self.Cardtype)
                self.carddict()
                
                print(self.NodeNo,self.SlotNo,self.sampleNo,self.Redu,self.NodeNo,self.SlotNo,self.sampleNo,self.Redu,self.NodeNo,self.SlotNo,self.sampleNo,self.Redu,self.NodeNo,self.SlotNo,self.sampleNo,self.Redu,)
                
                if self.sampleNo != 0:
                    self.sampleAddr()
                    self.targetaddr()
                    DifCol = abs(self.MbMinCol - self.MinCol)
                    DifRow = (self.MbMaxRow - self.MaxRow)
                    #print(DifCol)
                    #print(DifRow)
                          
                    for row in  self.ws.iter_cols(max_col=self.MaxCol, max_row=self.MaxRow,min_col=self.MinCol,min_row=self.MinRow):
                        for cell in  row:
                            #print(cell.value)
                            self.ws_charge.cell(cell.row+DifRow,cell.column+DifCol).value = cell.value
                            #print(cell.row+DifRow ,cell.column+DifCol)
                            self.ws_charge.cell(cell.row+DifRow,cell.column+DifCol).font = Font(name = cell.font.name,size =cell.font.sz)
                            self.ws_charge.cell(cell.row+DifRow,cell.column+DifCol).fill = PatternFill(patternType="solid", fgColor=cell.fill.fgColor, bgColor=cell.fill.bgColor)
                            
                            B1 = cell.border.left
                            B2 = cell.border.right
                            B3 = cell.border.top
                            B4 = cell.border.bottom
                            self.ws_charge.cell(cell.row+DifRow,cell.column+DifCol).border = Border(left=Side(B1.style,B1.color),
                                                                            right=Side(B2.style,B2.color),
                                                                            top=Side(B3.style,B3.color),
                                                                            bottom=Side(B4.style,B4.color))
                                         
                    self.wb.save('NodePlan.xlsx')   
            print("finish")
    
#主程序======================================================   

if __name__ == "__main__":
    #root = Tk()
    App()