#!/usr/bin/python
# -*- coding: UTF-8 -*-

import random
from openpyxl import Workbook

def generate(max,max_count):
    count = 0
    all_ressult = []
    while(count < max_count):
        A = random.randint(1, max)
        B = random.randint(1, max)
        symbol = random.choice('+-')
        result = ''

        if symbol == '+':
            if (A + B <max):
                result = str(A) + str(symbol) + str(B) + "="
                pass
                count = count + 1

        if symbol == '-':
            if (A - B > 0):
                result = str(A) + str(symbol) + str(B) + "="
                pass
                count = count + 1
        if result:
            #print(result,count)
            all_ressult.append(result)
    return all_ressult

def out():
    ressult = generate(20,40)
    wb = Workbook()
    sheet = wb.create_sheet('demo_sheet',index=0)
    for row in range(1, 11):
        for col in range(2, 6):
            index = (col-2)*10 + row -1
            sheet.cell(column=col*2, row=row, value="{0}".format(ressult[index]))
    wb.save('demo.xlsx')  # 保存文件，注意以xlsx为文件扩


if __name__ == "__main__":

    out()