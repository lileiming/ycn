# -*- coding:utf-8 -*-
#!/usr/bin/python
# python 3.8
# **********************************************************
# BK ENG Tool Trend = BKEToolTrend
# 趋势检查表格生成工具、趋势组态文件生成工具
# **********************************************************
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import os
import re
import YokoCustomlibrary   #自定义模块
from time import sleep
from YokoCustomlibrary import time_Decorator,thread_Decorator
import openpyxl
import openpyxl.styles as sty

class Windows_NODE(YokoCustomlibrary.FILE_NODE):

    def __init__(self, master):
        self.master = master
        self.here = os.getcwd()
        self.initWidgets()
        help_doc = '本程序为趋势检查表格生成工具、趋势组态文件生成工具 \n' \
                   '\n' \
                   'Trend_temp(位号列表) 转化 Checksheet(检查表)\n' \
                   'Checksheet(检查表) 转化 Csvsheet（组态文件）\n' \
                   '\n' \
                   '分成两步目的:\n' \
                   '如果用户返回Checksheet(检查表)，我们就执行第二步\n' \
                   '如果不提供就两步都执行！\n'
        self.text_update(help_doc)
    pass

    def initWidgets(self):
        test_path = 'C:/Users/Administrator/Documents/python/0.输出文件/ToolTrend/Trend_list.xlsx'
        # 创建中部
        mid_frame = LabelFrame(self.master, text='生成检查表格')
        mid_frame.pack(fill=X, padx=15, pady=0)
        self.e2 = StringVar()
        self.entry2 = ttk.Entry(mid_frame, width=45, textvariable=self.e2)
        self.e2.set(test_path)
        self.entry2.pack(fill=X, expand=YES, side=LEFT, pady=10)
        self.e3 = StringVar()
        self.comboxlist = ttk.Combobox(mid_frame, width=15, textvariable=self.e3)
        self.comboxlist.pack(side=LEFT)
        self.e3.set('Trend_temp')
        ttk.Button(mid_frame, text='数据列表', command=self.open_file2).pack(side=LEFT)

        mid_frame1 = LabelFrame(self.master, text='生成趋势文件')
        mid_frame1.pack(fill=X, padx=15, pady=0)
        self.e21 = StringVar()
        self.entry21 = ttk.Entry(mid_frame1, width=45, textvariable=self.e21)
        self.e21.set(test_path)
        self.entry21.pack(fill=X, expand=YES, side=LEFT, pady=10)
        self.e31 = StringVar()
        self.comboxlist1 = ttk.Combobox(mid_frame1, width=15, textvariable=self.e31)
        self.comboxlist1.pack(side=LEFT)
        self.e31.set("Checksheet")
        ttk.Button(mid_frame1, text='数据列表', command=self.open_file21).pack(side=LEFT)

        # 创建中下
        bot_frame1 = LabelFrame(self.master, text='结果')
        bot_frame1.pack(fill=X, side=TOP, padx=15, pady=0)
        self.Scroll = Scrollbar(bot_frame1)
        self.Text = Text(bot_frame1, width=83, height=13, yscrollcommand=self.Scroll.set)
        self.Text.pack(side=LEFT, padx=0, pady=5)
        self.Scroll = Scrollbar(bot_frame1)
        self.Scroll.pack(side=LEFT, fill=Y)
        self.Scroll.config(command=self.Text.yview)

        # 创建底部
        bot_frame = LabelFrame(self.master)
        bot_frame.pack(fill=X, side=TOP, padx=15, pady=8)
        self.e = StringVar()
        ttk.Label(bot_frame, width=40, textvariable=self.e).pack(side=LEFT, fill=BOTH, expand=YES, pady=10)
        self.e.set('程序员美德：懒惰、不耐烦、傲慢')
        ttk.Button(bot_frame, text='趋势文件', command=self.csv_command).pack(side=RIGHT, padx=10)
        ttk.Button(bot_frame, text='检查表格', command=self.command).pack(side=RIGHT, padx=10)
    pass

    def open_file2(self):
        file_path = filedialog.askopenfilename(title=u'选择数据列表', initialdir=self.here)
        file_text = file_path
        self.entry2.delete(0, END)
        self.entry2.insert('insert', file_text)
        sheetName = self.get_sheet(file_text)
        sheetNameT = tuple(sheetName)
        self.comboxlist["values"] = sheetNameT
        self.comboxlist.current(0)
    pass

    def open_file21(self):
        file_path = filedialog.askopenfilename(title=u'选择数据列表', initialdir=self.here)
        file_text = file_path
        self.entry21.delete(0, END)
        self.entry21.insert('insert', file_text)
        sheetName = self.get_sheet(file_text)
        sheetNameT = tuple(sheetName)
        self.comboxlist1["values"] = sheetNameT
        self.comboxlist1.current(0)
    pass

    @thread_Decorator
    @time_Decorator
    def command(self):
        #开始
        self.text_update('START_')
        self.read_date()
        self.write_date_check()
        self.text_update('STOP_')
        try:
            os.startfile(self.entry2.get())
        except FileNotFoundError:
            self.text_update('错误提示：文件不存在')
            pass
    pass

    def read_date(self):
        self.ditc_Trend = {}
        #读取界面 文件信息
        file_name = self.entry2.get()
        sheet_name = self.comboxlist.get()
        Keywords= list(next(self.get_data_Tag(file_name, sheet_name)))
        for i in self.get_data(file_name, sheet_name):
            tag_index = i[Keywords[0]]
            tag_key = i[Keywords[1]]
            tag_item = i[Keywords[2]]
            if tag_key != "":
                self.ditc_Trend[tag_index] = f'{tag_key}.{tag_item}'
            else:
                self.ditc_Trend[tag_index] = ""
            pass
        pass
    pass

    def write_date_check(self):
        # 3 Trend Group Checksheet
        column_ini = 2
        row_ini = 2
        file_name = self.entry2.get()
        data = openpyxl.load_workbook(file_name)
        #table = data.create_sheet("Csvsheet")
        table = data["Checksheet"]
        for i  in range(1,len(self.ditc_Trend)+1):
            trend_row = int(i / 8)
            trend_column = i % 8
            if trend_column == 0:
                trend_row = trend_row - 1
                trend_column = 8
            pass

            trend_Block = int(i/128)+1
            if i%128 == 0:
                trend_Block = int(i/128)
            pass
            rend_Block_str = str(trend_Block).zfill(2)

            trend_row = trend_row + row_ini
            trend_column = trend_column + column_ini

            trend_Group = (trend_row-1)%16
            if trend_Group == 0:
                trend_Group = 16
            pass
            trend_Group_str = str(trend_Group).zfill(2)

            #print(trend_row,trend_column)
            table.cell(row=trend_row, column=1).value = trend_Block
            table.cell(row=trend_row, column=2).value = trend_Group
            table.cell(row=trend_row, column=trend_column).value = self.ditc_Trend[i]
            table.cell(row=trend_row, column=11).value = f"TG{rend_Block_str}{trend_Group_str}"
            table.cell(row=trend_row, column=12).value = f"Pass / Fail"
        pass
        try:
            data.save(file_name)
        except PermissionError:
            self.text_update("请勿打开文件！\n")
        pass
    pass

    @thread_Decorator
    @time_Decorator
    def csv_command(self):
        #开始
        self.text_update('START_')
        self.read_date_csv()
        self.write_date_csv()
        self.text_update('STOP_')

        try:
            os.startfile(self.entry21.get())
        except FileNotFoundError:
            self.text_update('错误提示：文件不存在')
            pass
    pass

    def read_date_csv(self):
        self.ditc_csv = {}
        self.list_WindowName = []
        #读取界面 文件信息
        file_name = self.entry21.get()
        sheet_name = self.comboxlist1.get()
        Keywords= list(next(self.get_data_Tag(file_name, sheet_name)))
        for i in self.get_data(file_name, sheet_name):
            tag_Block = i[Keywords[0]]
            tag_Group = i[Keywords[1]]
            tag_TagName1 = i[Keywords[2]]
            tag_TagName2 = i[Keywords[3]]
            tag_TagName3 = i[Keywords[4]]
            tag_TagName4 = i[Keywords[5]]
            tag_TagName5 = i[Keywords[6]]
            tag_TagName6 = i[Keywords[7]]
            tag_TagName7 = i[Keywords[8]]
            tag_TagName8 = i[Keywords[9]]
            tag_Window_Name = i[Keywords[10]]
            tag_list = [tag_Block,tag_Group,tag_TagName1,tag_TagName2,tag_TagName3,tag_TagName4,tag_TagName5,tag_TagName6,tag_TagName7,tag_TagName8]
            self.ditc_csv[tag_Window_Name] = tag_list
            self.list_WindowName.append(tag_Window_Name)
        pass
    pass

    def write_date_csv(self):
        index_ini = 5
        file_name = self.entry21.get()
        data = openpyxl.load_workbook(file_name)
        #table = data.create_sheet("Csvsheet")
        table = data["Csvsheet"]
        for name in self.list_WindowName:
            index = self.list_WindowName.index(name)
            #print(index)
            for pen in range(1,9):
                index_row = index_ini+pen+index*8
                table.cell(row=index_row, column=2).value = pen
                table.cell(row=index_row, column=1).value = self.ditc_csv[name][1]
                table.cell(row=index_row, column=3).value = self.ditc_csv[name][pen + 1]

                if pen == 1 and self.ditc_csv[name][1] ==1:
                    table.cell(row=index_row, column=1).fill = sty.PatternFill(fill_type='solid',fgColor="FF00FF")  # 对更新数据进行标注颜色
                pass

                if self.ditc_csv[name][pen + 1] != "":
                    table.cell(row=index_row, column=4).value = 0
                    table.cell(row=index_row, column=7).value = 'Default'
                    table.cell(row=index_row, column=8).value = 1
                    table.cell(row=index_row, column=9).value = 1
                    table.cell(row=index_row, column=10).value = 1
                pass
            pass
        pass
        try:
            data.save(file_name)
        except PermissionError:
            self.text_update("请勿打开文件\n")
    pass

    def text_update(self,show):
        if show == 'START_':
            self.Text.insert(END, "=============程序开始=============\n")
        elif show == 'STOP_':
            self.Text.insert(END, "=============程序结束=============\n")
        else:
            self.Text.insert(END,show)
        self.Text.update()
        self.Text.see(END)
        #self.text_update('START_')
        pass
    pass

if __name__ == "__main__":
        root = Tk()
        root.title("V1.00")
        root.geometry('640x400')  # 窗口尺寸
        Windows_NODE(root)
        limit_time = YokoCustomlibrary.ALRM_NODE.limited_time(root)
        root.title("趋势生成工具" + "    到期日:" + limit_time)
        root.mainloop()
