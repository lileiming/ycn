# -*- coding:utf-8 -*-
#!/usr/bin/python
# python 3.8
# ==========================================================
# 参考文档：https://cloud.tencent.com/developer/section/1372347
#           http://c.biancheng.net/view/2539.html
# Rev01,Rev02
# 可以读取Yget 清单，并使用第四个表 的内容替换 原来清单数据 
# 只能更换一个数据

# Rev03
# 替换 多个数据

# Rev05
# 第3张表数据 更新第2张表内容
# 配合表格 根据Yget清单跑机柜接线表

# Rev06
# 增加一个配置文件
# 报警弹窗

# Rev07
# 增加 校队功能

# Rev08
# 增加更多参数
# 增加文档信息确认的弹窗

# Rev09
# 修改为界面模式

# Rev1.0
# 修正一些错误 加入限制
# Tag_list维护工具
# ==========================================================
import xlrd
import openpyxl
import os
import openpyxl.styles as sty
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from time import sleep
import YokoRead   #自定义模块
from YokoRead import time_Decorator,thread_Decorator
from sys import argv


# 计时装饰器

class Windows_NODE(YokoRead._FILE_NODE_):
    def __init__(self, master):
        self.master = master
        self.here = os.getcwd()
        self.initWidgets()
        help_doc = 'Don\'t Repeat Youself!!\n '
        self.text_update(help_doc)

    def initWidgets(self):
        # 创建第一层
        top_frame = LabelFrame(self.master, height=150, width=615, text="目标文件")
        top_frame.pack(side=TOP, padx=0, pady=5)
        # 目标文件名显示框
        self.tope1 = StringVar()
        self.entry = ttk.Entry(top_frame, width=85, textvariable=self.tope1)
        self.tope1.set("目标文件")
        self.entry.pack(fill=X, expand=0, side=TOP, padx=0, pady=10)
        # 目标文件Sheet框
        self.tope2 = StringVar()
        self.comboxlist = ttk.Combobox(top_frame, width=25, textvariable=self.tope2)
        self.comboxlist.pack(side=LEFT)
        self.comboxlist.bind("<<ComboboxSelected>>", self.open_combox)
        self.tope2.set("目标sheet")
        #  目标文件PID_TAG框
        self.tope3 = StringVar()
        self.comboxlist1 = ttk.Combobox(top_frame, width=25, textvariable=self.tope3)
        self.comboxlist1.pack(side=LEFT, padx=10, pady=10)
        self.tope3.set("目标TAG")
        ## 同文件选择框
        self.intVar = BooleanVar()
        # should_auto = BooleanVar()
        self.check1 = Checkbutton(top_frame, text="同文件", variable=self.intVar, command=self.change).pack(side=LEFT)
        # self.check1.select()
        self.ckeckchange = 0
        ## 目标文件选择按钮
        ttk.Button(top_frame, text='目标文件', command=self.open_file).pack(side=RIGHT, padx=5)

        # 创建第二层
        mid_frame1 = LabelFrame(self.master, height=150, width=615, text="参考文件")
        mid_frame1.pack(side=TOP, padx=0, pady=5)
        ## 参考文件名显示框
        self.mide1 = StringVar()
        self.entry2 = ttk.Entry(mid_frame1, width=85, textvariable=self.mide1)
        self.mide1.set("参考文件")
        self.entry2.pack(fill=X, expand=0, side=TOP, padx=0, pady=10)

        self.mide2 = StringVar()
        self.comboxlist2 = ttk.Combobox(mid_frame1, width=25, textvariable=self.mide2)
        self.comboxlist2.pack(side=LEFT)
        self.comboxlist2.bind("<<ComboboxSelected>>", self.open_combox2)
        self.mide2.set("参考sheet")
        self.mide3 = StringVar()
        self.comboxlist3 = ttk.Combobox(mid_frame1, width=25, textvariable=self.mide3)
        self.comboxlist3.pack(side=LEFT, padx=10, pady=10)
        self.mide3.set("参考TAG")
        ttk.Button(mid_frame1, text='参考文件', command=self.open_file2).pack(side=RIGHT, padx=5)

        # 创建第三层
        mid_frame2 = LabelFrame(self.master, height=60, width=615, text="结果")
        mid_frame2.pack(side=TOP, padx=0, pady=5)
        self.Scroll = Scrollbar(mid_frame2)
        self.Text = Text(mid_frame2, width=83, height=20, yscrollcommand=self.Scroll.set)
        self.Text.pack(side=LEFT, padx=0, pady=5)
        self.Scroll = Scrollbar(mid_frame2)
        self.Scroll.pack(side=LEFT, fill=Y)
        self.Scroll.config(command=self.Text.yview)

        # 创建底部部
        bot_frame = Frame(self.master, height=40, width=615)
        bot_frame.pack(side=TOP, padx=0, pady=5)
        ttk.Button(bot_frame, text='校队', command=self.ExcelComp).pack(side=LEFT, padx=10)
        ttk.Button(bot_frame, text='替换', command=self.ExcelSave).pack(side=LEFT, padx=10)
        ttk.Button(bot_frame, text='查看结果', command=self.OpenExcel).pack(side=LEFT, padx=10)

    @thread_Decorator
    def change(self):
        self.entry2.delete(0, END)
        self.ckeckchange = self.intVar.get()
        if (self.ckeckchange):
            self.open_file2()

    def open_file(self):
        try:
            file_path = filedialog.askopenfilename(title=u'选择参考文档', initialdir=self.here)
            file_text = file_path
            sheetName = self.get_sheet(file_text)
            self.sheetNameT = tuple(sheetName)
            self.comboxlist["values"] = self.sheetNameT
            self.comboxlist.current(0)

            self.entry.delete(0, END)
            self.entry.insert('insert', file_text)

        except xlrd.biffh.XLRDError as e:
            self.text_update('错误提示：文件格式错误，现在就只能处理Excel文档\n')

    def open_combox(self, *args):
        filename0 = self.entry.get()
        SheetName0 = self.comboxlist.get()
        self.col = self.get_col(filename0, SheetName0)
        self.comboxlist1["values"] = self.col
        self.comboxlist1.current(0)

    def open_file2(self):
        self.Text.delete(0.0, END)
        try:
            self.entry2.delete(0, END)
            if (self.ckeckchange):
                file_path = self.entry.get()
            else:
                file_path = filedialog.askopenfilename(title=u'选择数据列表', initialdir=self.here)
            file_text = file_path
            self.entry2.insert('insert', file_text)
            sheetName = self.get_sheet(file_text)
            self.sheetNameR = tuple(sheetName)
            self.comboxlist2["values"] = self.sheetNameR
            self.comboxlist2.current(0)
        except xlrd.biffh.XLRDError as e:
            self.text_update('错误提示：文件格式错误，现在就只能处理Excel文档\n')

    def open_combox2(self, event):
        filename = self.entry2.get()
        SheetName = self.comboxlist2.get()
        filename0 = self.entry.get()
        SheetName0 = self.comboxlist.get()

        self.col2 = self.get_col(filename, SheetName)
        self.comboxlist3["values"] = self.col2
        self.comboxlist3.current(0)

        if (filename == filename0 and SheetName == SheetName0):
            self.text_update("不是用重复文件\n")
        else:
            self.Text.delete(0.0, END)

    @thread_Decorator
    def ExcelSave(self):  # 替换功能
        try:
            self.Text.delete(0.0, END)
            self.text_update("=============替换中===============\n")
            self.Text.update()
            TExcleName = self.entry.get()
            self.TSheet = self.comboxlist.get()
            TSheetNum = self.sheetNameT.index(self.TSheet)
            # print (TSheetNum)
            Ttag = self.comboxlist1.get()
            TtagColNum = self.col.index(Ttag)
            # print (TtagColNum)
            RExcleName = self.entry2.get()
            RSheet = self.comboxlist2.get()
            RSheetNum = self.sheetNameR.index(RSheet)
            # print (RSheetNum)
            Rtag = self.comboxlist3.get()
            RtagColNum = self.col2.index(Rtag)
            # print (RtagColNum)

            TfilePath = os.path.join(os.getcwd(), TExcleName)
            TexcelFile = xlrd.open_workbook(TfilePath)
            RfilePath = os.path.join(os.getcwd(), RExcleName)
            RexcelFile = xlrd.open_workbook(RfilePath)
            # print excelFile.sheet_names()
            sheet1 = TexcelFile.sheet_by_index(TSheetNum)
            sheet2 = RexcelFile.sheet_by_index(RSheetNum)

            # ==========================================================
            rowsNum = sheet1.nrows
            colNum = sheet1.ncols
            # print "row num:", sheet1.nrows #行
            # print "col num:", sheet1.ncols #列
            rowlist_list = self.rowList(sheet1, TtagColNum)
            collist_list = self.colList(sheet1)
            dictTAG = self.dictGenerate(sheet1, TtagColNum)

            # 读取更换内容==============================================
            newTAG = self.dictGenerate(sheet2, RtagColNum)
            list_TAG = list(newTAG)
            # print list_TAG # [(u'BC', u'AB')]
            TAG_index = 0
            # openpyxl 打开Yget 清单=============================================
            data = openpyxl.load_workbook(TExcleName)
            table = data.worksheets[TSheetNum]

            while (TAG_index < len(list_TAG)):
                TAG_XY = list_TAG[TAG_index]
                # print TAG_XY
                TAG_Z = newTAG[TAG_XY]
                # print TAG_Z # 7.0
                Nokey = dictTAG.get(TAG_XY, 0)  # 处理Yget 不存在的key

                if (TAG_Z != "" and Nokey != 0):
                    # print(TAG_Z)
                    TAG_Z_OUT = str(TAG_Z) + '\n'
                    # Excel坐标============================================
                    NewTagRow_Str = list_TAG[TAG_index][0]
                    NewTagcol_Str = list_TAG[TAG_index][1]
                    NewTagRow_index = rowlist_list.index(NewTagRow_Str) + 1
                    NewTagCel_index = collist_list.index(NewTagcol_Str) + 1
                    # 替换Excel数据===============================================
                    TAG_ZZ = table.cell(NewTagRow_index, NewTagCel_index).value
                    if (TAG_Z != TAG_ZZ):
                        table.cell(NewTagRow_index, NewTagCel_index).value = TAG_Z
                        table.cell(NewTagRow_index, NewTagCel_index).fill = sty.PatternFill(fill_type='solid',
                                                                                            fgColor="00FFFF")  # 对更新数据进行标注颜色
                        self.text_update(TAG_Z_OUT)
                TAG_index = TAG_index + 1
                # Excel写入===========================================================
            data.save(TExcleName)
            self.text_update("=============替换结束=============\n")
        except Exception as e:
            self.text_update(e)
        sleep(2)

    @thread_Decorator
    @time_Decorator
    def ExcelComp(self):  # 比较功能
        try:
            #self.Text.delete(0.0, END)
            self.text_update("=============校队中===============\n")
            TExcleName = self.entry.get()
            self.TSheet = self.comboxlist.get()
            TSheetNum = self.sheetNameT.index(self.TSheet)
            # print (TSheetNum)
            Ttag = self.comboxlist1.get()
            TtagColNum = self.col.index(Ttag)
            # print (TtagColNum)
            RExcleName = self.entry2.get()
            RSheet = self.comboxlist2.get()
            RSheetNum = self.sheetNameR.index(RSheet)
            # print (RSheetNum)
            Rtag = self.comboxlist3.get()
            RtagColNum = self.col2.index(Rtag)
            # print (RtagColNum)

            TfilePath = os.path.join(os.getcwd(), TExcleName)
            TexcelFile = xlrd.open_workbook(TfilePath)
            RfilePath = os.path.join(os.getcwd(), RExcleName)
            RexcelFile = xlrd.open_workbook(RfilePath)
            # print excelFile.sheet_names()

            sheet1 = TexcelFile.sheet_by_index(TSheetNum)
            sheet2 = RexcelFile.sheet_by_index(RSheetNum)

            # ==========================================================
            rowsNum = sheet1.nrows
            colNum = sheet1.ncols
            # print "row num:", sheet1.nrows #行
            # print "col num:", sheet1.ncols #列
            rowlist_list = self.rowList(sheet1, TtagColNum)
            collist_list = self.colList(sheet1)
            dictTAG = self.dictGenerate(sheet1, TtagColNum)

            # 读取更换内容==============================================
            newTAG = self.dictGenerate(sheet2, RtagColNum)
            list_TAG = list(newTAG)
            # print list_TAG # [(u'BC', u'AB')]
            TAG_index = 0
            # openpyxl 打开Yget 清单=============================================
            data = openpyxl.load_workbook(TExcleName)
            table = data.worksheets[TSheetNum]

            while (TAG_index < len(list_TAG)):
                TAG_XY = list_TAG[TAG_index]
                # print TAG_XY
                TAG_Z = newTAG[TAG_XY]
                # print TAG_Z # 7.0
                Nokey = dictTAG.get(TAG_XY, 0)  # 处理Yget 不存在的key
                if (TAG_Z != "" and Nokey != 0):

                    # Excel坐标============================================
                    NewTagRow_Str = list_TAG[TAG_index][0]
                    NewTagcol_Str = list_TAG[TAG_index][1]
                    NewTagRow_index = rowlist_list.index(NewTagRow_Str) + 1
                    NewTagCel_index = collist_list.index(NewTagcol_Str) + 1
                    # 替换Excel数据===============================================
                    TAG_ZZ = table.cell(NewTagRow_index, NewTagCel_index).value
                    # table.cell(NewTagRow_index,NewTagCel_index).value = TAG_Z
                    if (TAG_Z != TAG_ZZ):
                        # print (TAG_Z)
                        TAG_Z_OUT = str(TAG_Z) + '\n'
                        self.Text.insert('insert', TAG_Z_OUT)
                        self.Text.update()
                        self.Text.see(END)
                        table.cell(NewTagRow_index, NewTagCel_index).fill = sty.PatternFill(fill_type='solid',
                                                                                            fgColor="FF00FF")  # 对更新数据进行标注颜色
                TAG_index = TAG_index + 1
                # Excel写入===========================================================
            data.save(TExcleName)
            self.text_update("=============校队结束=============\n")
        except Exception as e:
            self.text_update(e)
        sleep(2)

    def get_col(self, filename, SheetName):
        dir_case = filename
        data = xlrd.open_workbook(dir_case)
        sheet = data.sheet_by_name(SheetName)
        col_name = self.colList(sheet)
        return col_name

    def OpenExcel(self):
        try:
            os.startfile(self.entry.get())
        except FileNotFoundError:
            self.text_update('错误提示：文件不存在')

            pass

    def text_update(self,show):
        if show == 'START_':
            self.Text.insert(END, "=============程序开始=============\n")
        elif show == 'STOP_':
            self.Text.insert(END, "=============程序结束=============\n")
        else:
            self.Text.insert(END,show)
            pass
        self.Text.update()
        self.Text.see(END)
        #self.text_update('STOP_')
        pass

if __name__ == "__main__":
    try:
        str_ = argv[1]
    except:
        str_ = ''
    root = Tk()
    root.geometry('640x600+100+200')  # 窗口尺寸
    Windows_NODE(root)
    # limit_time = YokoRead._ALRM_NODE_.limited_time(root)
    if str_ != '-S':
        limit_time = YokoRead._ALRM_NODE_.limited_time(root)
    else:
        limit_time = '3020/1/1'
    root.title("Tag_list维护工具  Ver1.0" + "    到期日:" + limit_time)
    root.mainloop()
    pass
