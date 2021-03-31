# -*- coding:utf-8 -*-
#!/usr/bin/python
# python 3.8
# ==========================================================
# 参考文档：https://cloud.tencent.com/developer/section/1372347
#           http://c.biancheng.net/view/2539.html

# **********************************************************
# Excel Tool Tag list  = ExcelToolTaglist
# Tag_list维护工具
# **********************************************************

import xlrd
import openpyxl
import pandas as pd
import os
import openpyxl.styles as sty
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from time import sleep
import YokoCustomlibrary   #自定义模块
from YokoCustomlibrary import time_Decorator,thread_Decorator
from math import isnan
import configparser


# 计时装饰器

class Windows_NODE:
    def __init__(self, master):
        self.master = master

    def initWidgets(self):
        # 创建第一层
        top_frame = LabelFrame(self.master, height=150, width=615, text="目标文件")
        top_frame.pack(side=TOP, padx=0, pady=5)
        # 目标文件名显示框
        self.tope1 = StringVar()
        self.entry = ttk.Entry(top_frame, width=85, textvariable=self.tope1)
        self.tope1.set(self.var_ini_tope1)
        #self.tope1.set("目标文件")

        self.entry.pack(fill=X, expand=0, side=TOP, padx=0, pady=10)
        # 目标文件Sheet框
        self.tope2 = StringVar()
        self.comboxlist = ttk.Combobox(top_frame, width=25, textvariable=self.tope2)
        self.comboxlist.pack(side=LEFT)
        self.comboxlist.bind("<<ComboboxSelected>>", self.open_combox)
        self.tope2.set(self.var_ini_tope2)
        #  目标文件PID_TAG框
        self.tope3 = StringVar()
        self.comboxlist1 = ttk.Combobox(top_frame, width=20, textvariable=self.tope3)
        self.comboxlist1.pack(side=LEFT, padx=10, pady=10)
        self.tope3.set(self.var_ini_tope3)
        #  目标文件起始行
        self.tope4 = StringVar()
        self.entry3 = ttk.Entry(top_frame, width=5, textvariable=self.tope4)
        self.tope4.set(self.var_ini_tope4)
        self.entry3.pack(side=LEFT, pady=10)
        ## 同文件选择框
        self.intVar = BooleanVar()
        # should_auto = BooleanVar()
        self.check1 = Checkbutton(top_frame, text="同文件", variable=self.intVar, command=self.change).pack(side=LEFT)
        # self.check1.select()
        self.ckeckchange = 0
        ## 目标文件选择按钮
        ttk.Button(top_frame, text='目标文件', command= self.open_file).pack(side=RIGHT, padx=5)

        # 创建第二层
        mid_frame1 = LabelFrame(self.master, height=150, width=615, text="参考文件")
        mid_frame1.pack(side=TOP, padx=0, pady=5)
        ## 参考文件名显示框
        self.mide1 = StringVar()
        self.entry2 = ttk.Entry(mid_frame1, width=85, textvariable=self.mide1)
        self.mide1.set(self.var_ini_mide1)
        self.entry2.pack(fill=X, expand=0, side=TOP, padx=0, pady=10)

        self.mide2 = StringVar()
        self.comboxlist2 = ttk.Combobox(mid_frame1, width=25, textvariable=self.mide2)
        self.comboxlist2.pack(side=LEFT)
        self.comboxlist2.bind("<<ComboboxSelected>>", self.open_combox2)
        self.mide2.set(self.var_ini_mide2)
        self.mide3 = StringVar()
        self.comboxlist3 = ttk.Combobox(mid_frame1, width=20, textvariable=self.mide3)
        self.comboxlist3.pack(side=LEFT, padx=10, pady=10)
        self.mide3.set(self.var_ini_mide3)
        #  参考文件起始行
        self.mide4 = StringVar()
        self.entry4 = ttk.Entry(mid_frame1, width=5, textvariable=self.mide4)
        self.mide4.set(self.var_ini_mide4)
        self.entry4.pack(side=LEFT, pady=10)

        ## 空数据有效
        self.intVar_nan = BooleanVar()
        self.check_nan = Checkbutton(mid_frame1, text="空数据有效", variable=self.intVar_nan).pack(side=LEFT)
        #self.intVar_nan.get()

        ttk.Button(mid_frame1, text='参考文件', command=self.open_file2).pack(side=RIGHT, padx=5)

        # 创建第三层
        mid_frame2 = LabelFrame(self.master, height=60, width=615, text="结果")
        mid_frame2.pack(side=TOP, padx=0, pady=5)
        self.Scroll = Scrollbar(mid_frame2)
        self.Text = Text(mid_frame2, width=83, height=20, yscrollcommand=self.Scroll.set)
        self.Text.pack(side=LEFT, padx=0, pady=5)
        self.Scroll = Scrollbar(mid_frame2)
        self.Scroll.pack(side=LEFT, fill=Y)
        self.Scroll.config(command=self.Text.yview)

        # 创建底部部
        bot_frame = Frame(self.master, height=40, width=615)
        bot_frame.pack(side=TOP, padx=0, pady=5)
        ttk.Button(bot_frame, text='量程', command=self.Excel_Chang_Rang).pack(side=LEFT, padx=10)
        ttk.Button(bot_frame, text='校队', command=lambda:self.ExcelSave(1)).pack(side=LEFT, padx=10)
        ttk.Button(bot_frame, text='替换', command=lambda:self.ExcelSave(0)).pack(side=LEFT, padx=10)
        ttk.Button(bot_frame, text='查看结果', command=self.OpenExcel).pack(side=LEFT, padx=10)

class Function_NODE:
    def __init__(self):
        #self.here = os.getcwd()
        self.initpath('read')
        self.here = self.var_ini_here
        self.initWidgets()
        help_doc = 'Don\'t Repeat Youself!!\n ' \
                   '2021.1 新增：AD清单量程保留4有效数字。\n '\
                   '2021.2 新增：初始化单元。 \n' \

        self.text_update(help_doc)
        self.fun_init_window()

    @thread_Decorator
    def change(self):
        self.entry2.delete(0, END)
        self.ckeckchange = self.intVar.get()
        if self.ckeckchange:
            self.open_file2()

    def open_file(self):
        try:
            var_file_path = filedialog.askopenfilename(title=u'选择参考文档', initialdir=self.here)
            list_sheet_name = pd.ExcelFile(var_file_path).sheet_names  # 读取文件的所有表单名，得到列表
            self.sheet_name_T = tuple(list_sheet_name)
            self.comboxlist["values"] = self.sheet_name_T
            self.comboxlist.current(0)
            self.entry.delete(0, END)
            self.entry.insert('insert', var_file_path)
            self.var_header_T = int(self.entry3.get())

        except xlrd.biffh.XLRDError:
            self.text_update('错误提示：文件格式错误，现在就只能处理Excel文档\n')
        except FileNotFoundError:
            pass

    def open_combox(self, *args):
        self.var_header_T = int(self.entry3.get())
        var_file_path = self.entry.get()
        var_sheet_name = self.comboxlist.get()
        self.var_col = list(pd.ExcelFile(var_file_path).parse(var_sheet_name,self.var_header_T).keys())
        #self.var_col = [element for element in self.var_col if not "Unnamed" in element]
        self.comboxlist1["values"] = self.var_col
        self.comboxlist1.current(0)

    def open_file2(self):
        self.Text.delete(0.0, END)
        try:
            self.entry2.delete(0, END)
            if self.ckeckchange:
                var_file_path = self.entry.get()
            else:
                var_file_path = filedialog.askopenfilename(title=u'选择数据列表', initialdir=self.here)
            pass
            self.entry2.insert('insert', var_file_path)
            list_sheet_name = pd.ExcelFile(var_file_path).sheet_names
            self.sheet_name_R = tuple(list_sheet_name )
            self.comboxlist2["values"] = self.sheet_name_R
            self.comboxlist2.current(0)
        except xlrd.biffh.XLRDError:
            self.text_update('错误提示：文件格式错误，现在就只能处理Excel文档\n')
        except FileNotFoundError:
            pass

    def open_combox2(self,*args):
        self.var_header_R = int(self.entry4.get())
        var_file = self.entry2.get()
        var_sheet = self.comboxlist2.get()
        var_file_0 = self.entry.get()
        var_sheet_0 = self.comboxlist.get()
        self.var_col2 = list(pd.ExcelFile(var_file).parse(var_sheet,self.var_header_R).keys())
        #self.var_col2 = [element for element in self.var_col2 if not "Unnamed" in element]
        self.comboxlist3["values"] = self.var_col2
        self.comboxlist3.current(0)
        if var_file == var_file_0 and var_sheet == var_sheet_0:
            self.text_update("不是用重复文件\n")
        else:
            self.Text.delete(0.0, END)

    def fun_init_window(self):
        try:
            list_sheet_name = pd.ExcelFile(self.var_ini_tope1).sheet_names  # 读取文件的所有表单名，得到列表
            self.sheet_name_T = tuple(list_sheet_name)
            list_sheet_name1 = pd.ExcelFile(self.var_ini_mide1).sheet_names
            self.sheet_name_R = tuple(list_sheet_name1)

            self.var_header_T = int(self.var_ini_tope4)
            self.var_header_R = int(self.var_ini_mide4)

            self.var_col = list(pd.ExcelFile(self.var_ini_tope1).parse(self.var_ini_tope2, self.var_header_T).keys())
            self.var_col2 = list(pd.ExcelFile(self.var_ini_mide1).parse(self.var_ini_mide2, self.var_header_R).keys())

        except :
            print("初始化失败！")
            if not os.path.exists('ExcelToolTaglistConfig.ini'):
                file = open('ExcelToolTaglistConfig.ini', 'w')
                file.close()
                out_result = f'[section_1]'
                with open('ExcelToolTaglistConfig.ini', 'w', encoding='GBK') as out_file:  # 保存为ANSI格式
                    out_file.write(out_result)

    @thread_Decorator
    def ExcelSave(self,Comp):  # 替换功能
        try:
            self.Text.delete(0.0, END)
            if Comp == 0:
                self.text_update("=============替换中=============\n")
            if Comp == 1:
                self.text_update("=============校队中=============\n")
            T_Excle_Name = self.entry.get()
            self.TSheet = self.comboxlist.get()
            T_Sheet_Num = self.sheet_name_T.index(self.TSheet)
                                                                                                                        #从UI界面读取 目标文件信息
            R_Excle_Name = self.entry2.get()
            self.RSheet = self.comboxlist2.get()
            R_Sheet_Num = self.sheet_name_R.index(self.RSheet)
            R_Tag = self.comboxlist3.get()
                                                                                                                        #从UI界面读取 参考文件信息
            #=================================
            DataFrame_T = pd.read_excel(T_Excle_Name, sheet_name = T_Sheet_Num, header = self.var_header_T)             #目标文件 EXCLE to DataFrame
            DataFrame_R = pd.read_excel(R_Excle_Name, sheet_name = R_Sheet_Num, header = self.var_header_R)             #参考文件 EXCLE to DataFrame
            op_data = openpyxl.load_workbook(T_Excle_Name)
            op_table = op_data.worksheets[T_Sheet_Num]

            DataFrame_Temp = DataFrame_R.append(DataFrame_T)
            DataFrame_Temp = DataFrame_Temp.append(DataFrame_T)                                                         #两次导入 目标Sheet 再去重 就会留下需要更新内容
            #print(self.var_col2)                                                                                       #参考Sheet keys()
            DataFrame_Diff = DataFrame_Temp.drop_duplicates(subset=self.var_col2, keep=False)                           #去重
            DataFrame_Diff.reset_index(inplace=True,drop=True)                                                          #重置index
            #print(DataFrame_Diff)
            dict_Diff = DataFrame_Diff.to_dict(orient='index')                                                          #DataFrame 转 dict
            #print(dict_Diff)
            var_diff_len = len(dict_Diff)                                                                               #更新内容字典长度
            #print(var_diff_len)
            index_bias = 2
            col_bias = 1
            for diff_i in range(var_diff_len):
                Var_diff_Tag = dict_Diff[diff_i][R_Tag]                                                                 #读取更新内容TAG
                index_temp = DataFrame_T[(DataFrame_T[R_Tag] == Var_diff_Tag)].index.tolist()                           #读取更新内容TAG 在目标文件中定位 index
                #print(index_temp)
                if len(index_temp) != 0:
                    for col_i in range(1, len(self.var_col2)):
                        if self.var_col.__contains__(self.var_col2[col_i]):
                            var_old_data = DataFrame_T.at[index_temp[0], self.var_col2[col_i]]                          # 目标文件 原来数据
                            var_new_data = dict_Diff[diff_i][self.var_col2[col_i]]                                      # 更新内容 新数据
                            if var_old_data != var_new_data and str(var_new_data) != 'nan' :                            # 两者有区别 再更新
                                col_temp = self.var_col.index(self.var_col2[col_i])                                     #定位
                                if Comp == 0 :
                                    op_table.cell(index_temp[0] + index_bias + self.var_header_T, col_temp + col_bias).value = var_new_data
                                    op_table.cell(index_temp[0] + index_bias+ self.var_header_T, col_temp + col_bias).fill = sty.PatternFill(fill_type='solid',
                                                                                                    fgColor="00FFFF")       # 对更新数据进行标注颜色
                                if Comp == 1:
                                    op_table.cell(index_temp[0] + index_bias + self.var_header_T, col_temp + col_bias).fill = sty.PatternFill(fill_type='solid',
                                                                                                fgColor="FF00FF")           # 对更新数据进行标注颜色
                                self.text_update(f'{Var_diff_Tag}===={var_new_data}' + '\n')
                            pass
                        pass
                    pass
                pass
            pass
            op_data.save(T_Excle_Name)
            if Comp == 0:
                self.text_update("=============替换结束=============\n")
            if Comp == 1:
                self.text_update("=============校队结束=============\n")
            self.initpath('write')
        except Exception as e:
            self.text_update(e)
        sleep(1)

    @thread_Decorator
    def Excel_Chang_Rang(self):  # 量程
        try:
            self.text_update("=============量程修订中=============\n")
            self.Text.delete(0.0, END)
            T_Excle_Name = self.entry.get()
            self.TSheet = self.comboxlist.get()
            T_Sheet_Num = self.sheet_name_T.index(self.TSheet)
            R_Tag = self.comboxlist1.get()
            # 从UI界面读取 参考文件信息
            # =================================
            DataFrame_T = pd.read_excel(T_Excle_Name, sheet_name=T_Sheet_Num,
                                        header=self.var_header_T)                                                       # 目标文件 EXCLE to DataFrame
            op_data = openpyxl.load_workbook(T_Excle_Name)
            op_table = op_data.worksheets[T_Sheet_Num]
            Copy_DF= DataFrame_T.copy()
            if 'Range HI' in self.var_col:
                for i in range(len(Copy_DF['Range HI'])):
                    if isnan(Copy_DF['Range HI'][i]):
                        continue
                    pass
                    if not isnan(Copy_DF['Range HI'][i]):
                        list_hi_lo = self.func_Rang_HI(Copy_DF['Range LO'][i], Copy_DF['Range HI'][i])
                        Copy_DF.loc[i, 'Range LO'] = list_hi_lo[0]
                        Copy_DF.loc[i, 'Range HI'] = list_hi_lo[1]
                    pass
                pass
                dict_Diff = Copy_DF.to_dict(orient='index')
                var_diff_len = len(dict_Diff)                                                                           # 更新内容字典长度
                # print(var_diff_len)
                index_bias = 2
                col_bias = 1
                for diff_i in range(var_diff_len):
                    Var_diff_Tag = dict_Diff[diff_i][R_Tag]                                                             # 读取更新内容TAG
                    index_temp = DataFrame_T[(DataFrame_T[R_Tag] == Var_diff_Tag)].index.tolist()                       # 读取更新内容TAG 在目标文件中定位 index
                    # print(index_temp)
                    if len(index_temp) != 0:
                        for col_i in range(1, len(self.var_col)):
                            if self.var_col[col_i] == 'Range HI' or self.var_col[col_i] == 'Range LO':
                                var_old_data = DataFrame_T.at[index_temp[0], self.var_col[col_i]]                       # 目标文件 原来数据
                                var_new_data = dict_Diff[diff_i][self.var_col[col_i]]                                   # 更新内容 新数据
                                if var_old_data != var_new_data and str(var_new_data) != 'nan':                         # 两者有区别 再更新
                                    col_temp = self.var_col.index(self.var_col[col_i])                                  # 定位
                                    op_table.cell(index_temp[0] + index_bias + self.var_header_T,
                                                      col_temp + col_bias).value = var_new_data
                                    op_table.cell(index_temp[0] + index_bias + self.var_header_T,
                                                      col_temp + col_bias).fill = sty.PatternFill(fill_type='solid',
                                                                                                  fgColor="00FFFF")     # 对更新数据进行标注颜色
                                    self.text_update(f'{Var_diff_Tag}===={var_new_data}' + '\n')
                                pass
                            pass
                        pass
                    pass
                op_data.save(T_Excle_Name)
                pass
            pass
            self.text_update("=============量程修订结束=============\n")
            self.initpath('write')

        except Exception as e:
            self.text_update(e)
        sleep(1)

    def OpenExcel(self):
        try:
            os.startfile(self.entry.get())
        except FileNotFoundError:
            self.text_update('错误提示：文件不存在')
            pass

    def text_update(self,show):
        if show == 'START_':
            self.Text.insert(END, "=============程序开始=============\n")
        elif show == 'STOP_':
            self.Text.insert(END, "=============程序结束=============\n")
        else:
            self.Text.insert(END,show)
            pass
        self.Text.update()
        self.Text.see(END)
        #self.text_update('STOP_')
        pass

    def  initpath(self,parameter):
        try:
            self.cf = configparser.ConfigParser()
            self.cf.read("ExcelToolTaglistConfig.ini")

            if parameter == "read":
                self.var_ini_here = self.cf.get('section_1', 'here')
                self.var_ini_tope1 = self.cf.get('section_1', 'tope1')
                self.var_ini_tope2 = self.cf.get('section_1', 'tope2')
                self.var_ini_tope3 = self.cf.get('section_1', 'tope3')
                self.var_ini_tope4 = self.cf.get('section_1', 'tope4')
                self.var_ini_mide1 = self.cf.get('section_1', 'mide1')
                self.var_ini_mide2 = self.cf.get('section_1', 'mide2')
                self.var_ini_mide3 = self.cf.get('section_1', 'mide3')
                self.var_ini_mide4 = self.cf.get('section_1', 'mide4')

            if parameter == "write":
                path = os.path.dirname(self.entry.get()) #文件夹
                self.cf.set('section_1', 'here', path)
                self.cf.set('section_1', 'tope1', self.entry.get())
                self.cf.set('section_1', 'tope2', self.comboxlist.get())
                self.cf.set('section_1', 'tope3', self.comboxlist1.get())
                self.cf.set('section_1', 'tope4', self.entry3.get())
                self.cf.set('section_1', 'mide1', self.entry2.get())
                self.cf.set('section_1', 'mide2', self.comboxlist2.get())
                self.cf.set('section_1', 'mide3', self.comboxlist3.get())
                self.cf.set('section_1', 'mide4', self.entry4.get())
                self.cf.write(open('ExcelToolTaglistConfig.ini', 'w'))

        except:
            self.var_ini_here = os.getcwd()
            self.var_ini_tope1 = "目标文件"
            self.var_ini_tope2 = "Sheet"
            self.var_ini_tope3 = "PID_TAG"
            self.var_ini_tope4 = "0"
            self.var_ini_mide1 = "参考文件"
            self.var_ini_mide2 = "Sheet"
            self.var_ini_mide3 = "PID_TAG"
            self.var_ini_mide4 = "0"
        pass

    def func_Rang_HI(self,lower_input, upper_input):
        # input = 0.0
        out_put = []
        upper_input1 = float(upper_input)
        lower_input1 = float(lower_input)

        if upper_input1 < 10:
            out_put.append("{:.3f}".format(lower_input1))
            out_put.append("{:.3f}".format(upper_input1))
        if 10 <= upper_input1 < 100:
            out_put.append("{:.2f}".format(lower_input1))
            out_put.append("{:.2f}".format(upper_input1))
        if 100 <= upper_input1:
            out_put.append("{:.1f}".format(lower_input1))
            out_put.append("{:.1f}".format(upper_input1))
        return out_put

class SHOW_NODE(Windows_NODE,Function_NODE,YokoCustomlibrary.FILE_NODE):
    """
    多类继承
    """
    def __init__(self,master):
        Windows_NODE.__init__(self, master)
        Function_NODE.__init__(self)
    pass

if __name__ == "__main__":
    root = Tk()
    root.geometry('640x600+100+200')  # 窗口尺寸
    SHOW_NODE(root)
    limit_time = YokoCustomlibrary.ALRM_NODE.limited_time(root)
    root.title("Tag_list维护工具  Ver1.1" + "    到期日:" + limit_time)
    root.mainloop()
    pass

# 原因是最近xlrd更新到了2.0.1版本，只支持.xls文件。所以pandas.read_excel(‘xxx.xlsx’)会报错。
# 可以安装旧版xlrd，在cmd中运行：
# pip uninstall xlrd
# pip install xlrd==1.2.0