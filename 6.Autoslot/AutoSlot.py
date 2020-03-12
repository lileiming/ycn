# -*- coding:utf-8 -*-
#!/usr/bin/python
# python 3.7
#==========================================================
# Script Name	:  auto assign hardware
# Author		: lileiming@me.com
# Created		: 2019/11/4
# Last Modified	: 2019/11/11
# Version			: 1.0

# Modifications	: 0.1 frist version
# Modifications	: 1.0 基本实现卡件排布

# Description		: Automatically assign hardware locations based on big data.
#==========================================================

import pandas as pd
import math
from openpyxl import load_workbook
#import xlrd
#import xlwt
import time
import tkinter.messagebox

class App():
    def __init__(self):
        self.Excelfile = 'Sample.xlsx'                                          #定义文件名
        self.data = pd.read_excel(self.Excelfile,'Sheet0')                      #读取样本数据
        self.MaxIndex = len(self.data.index)                                    #数据索引最大值 14
        self.NodeSlotList = [[] for col in range(80)]                           #所有样本数据存取 2纬 List
        self.NodeSlotCal =  [[] for col in range(80)]                           #所有样本数据比例 2纬 List
        self.Card = []                                                          #需要排布的卡件列表
        self.outSheet = 'Out'

    def my_card_num(self):
        #需要处理卡件
        total = 0
        self.rowName = list(self.data.Nodeslot.unique())                        #需要排布的卡件名目
        self.cardTypeNum = len(self.rowName) - 1                                #需要排布的卡件类型数量

        for _ in range(self.cardTypeNum):
            cardtpye = self.data.iloc[_, 1]
            cardnum  = self.data.iloc[_, 2]
            for _ in range(int(cardnum)):
                self.Card.append(cardtpye)
                total = total + 1

        self.nodeNum = math.ceil (total / 8)                                    #需要node 数量
        self.total = int(self.nodeNum * 8)                                      #所有空插槽数量
        #print(self.total)
        #print(self.Card)
        for i in range(self.total):                                             #valid 行 标记为 -1 表示允许排卡
            self.data.iloc[self.cardTypeNum, i+2] = -1
        #print(self.data)

    def card(self):
        #排卡程序，
        self.colName = list(self.data.columns.values)
        #print(self.cardTypeNum)
        for Y in range(self.cardTypeNum):
            if (self.Card != []):
                for Z in range(4):                                              # 0~4
                    for X in range(self.nodeNum):                               # Node = 1-10
                        for W in range(2):                                      # 0~1
                            Q = 8*X+2*Z+W+2                                     #先扫描 S1&S2  在扫 S3&S4
                            #print(Q)
                            if (self.data.iloc[self.cardTypeNum,Q] == -1):
                                self.data = self.data.sort_values(by=str(self.colName[Q]), ascending=False)
                                #print(self.colName[X + 2])
                                nowCard = self.data.iloc[Y, 1]

                                if nowCard in self.Card:
                                    nowIndex = self.Card.index(nowCard)
                                    del self.Card[nowIndex]                     #排完的卡件，从list里删除

                                    if (nowCard != 'ADCV01'):                   #先拍非ADCV01卡
                                        self.data.iloc[self.cardTypeNum, Q] = nowCard
                                        print(nowCard)
                        self.re_queue()                                          #按照序号排一下
        #print(self.Card)
        for X in range(self.total):                                             #为拍卡的位置 排ADCV01
            if (self.data.iloc[self.cardTypeNum, X] == -1):
                self.data.iloc[self.cardTypeNum, X] = 'ADCV01'
                print('ADCV01')

    def re_queue(self):
        # 按照序号排序
        self.data = self.data.sort_values(by='NO.')
        self.data.index = range(self.MaxIndex)

    def read_data(self,sheetName):
        #读取 单张 Sheet 卡件布置图
        data = pd.read_excel(self.Excelfile, sheetName)
        for X in range(5):
            for Y in range(8):
                dataF = data.iloc[5*(X+1), (24 + Y)]                            # F面数据
                self.NodeSlotList[X * 8 + Y].append(dataF)
                #print(dataF)
        for X in range(5):
            for Y in range(8):
                dataR = data.iloc[5*(X+1), (42 + Y)]                            # R面数据
                self.NodeSlotList[X * 8 + Y + 40].append(dataR)
                #print(dataR)
        #print(self.NodeSlotList)

    def read_all_data(self):
        # 读取 第一张 Sheet 后所有的 Sheet
        for X in self.sheet_name_list()[1:]:                                      #读取第2张 Sheet 开始的数据。
            if self.outSheet not in X :
                self.read_data(X)
            if self.outSheet in X :
                #print(X)
                self.del_out_sheet(X)

    def pro_calcu(self):
        #Proportional calculation 比例计算
        Sum = 0.0
        #self.rowName = list(self.data.Nodeslot.unique())
        #print(self.colName)
        for Y in range(self.total):
            Sum = 0.0
            len1 = len(self.NodeSlotList[Y])
            #print(len1)
            for X in self.rowName[0:self.MaxIndex-2]:
                #print(X)
                count1 = self.NodeSlotList[Y].count(X)
                Cal = ((count1 / len1) * 0.98 + 0.001)* 100
                Sum = Sum + Cal
                self.NodeSlotCal[Y].append(Cal)
            self.NodeSlotCal[Y].append(100 - Sum)
            #print(self.NodeSlotCal[Y])
        for Y in range(self.cardTypeNum):
            for X in range(self.total):
                self.data.iloc[(Y), (X+2)] = self.NodeSlotCal[X][Y]

    def del_out_sheet(self,outSheeet):
        try:
            book = load_workbook(self.Excelfile)
            #print(book[outSheeet])
            book.remove(book[outSheeet])
            book.save(self.Excelfile)
            book.close()
        except PermissionError as e:
            print("输出文件未关闭")

    def to_file(self):
        #输出文件
        try:
            writer = pd.ExcelWriter(self.Excelfile)
            book = load_workbook(self.Excelfile)
            writer.book = book
            self.data.to_excel(writer, self.outSheet)
            #writer.save()
            writer.close()
        except PermissionError as e:
            print("输出文件未关闭")

    def sheet_name_list(self):
        #读取Sheet名
        self.SheetnameDict = pd.read_excel(self.Excelfile,None)
        return list(self.SheetnameDict.keys())

    def test(self):
        print(self.data.Nodeslot.unique())
        #print(self.data.iloc[13, 80])
        #print(self.NodeSlotCal[79][11])
        #self.data = self.data.sort_values(by='N1S1',ascending=False)
        #print(self.data)
        #self.data.iloc[12,1] = self.data.iloc[0,0]
        #print(self.data.iloc[12,1])

def limited_time():
    ticks = time.time()
    #print(ticks)
    limitTime = 1572414548+2592000
    #print(limitTime)
    localtime = time.strftime("%Y/%m/%d", time.localtime(limitTime))
    if (ticks > limitTime):
        tkinter.messagebox.showinfo("提示", "软件过期需要重新编译"+localtime)
        exit()

if __name__ == '__main__':
    #limited_time()
    app = App()
    app.my_card_num()
    #app.test()
    app.read_all_data()
    app.pro_calcu()
    app.card()
    #app.Requeue()
    app.to_file()


